#!/usr/bin/env python3
"""Generate test Excel file for teacher import with various error cases."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Template"

# Header style
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center")

# Headers (10 columns)
headers = [
    "NIP", "NUPTK", "Nama Lengkap", "Jenis Kelamin (male/female)",
    "Alamat", "No. HP", "Email", "Status Kepegawaian", "Jabatan", "Status (active/inactive)"
]

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

# Column widths
widths = [22, 22, 28, 24, 30, 18, 30, 20, 22, 18]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# Test data rows (mix of valid and invalid)
test_data = [
    # Row 2: VALID - Complete valid teacher
    ["198501012010011001", "2015123456700001", "Dr. Ahmad Fauzi, M.Pd", "male",
     "Jl. Pendidikan No. 10", "081234567890", "ahmad.fauzi@sekolah.sch.id",
     "PNS", "Guru Matematika", "active"],

    # Row 3: VALID - Female teacher, minimal data
    ["", "", "Siti Nurhaliza, S.Pd", "female",
     "", "081234567891", "siti.nurhaliza@sekolah.sch.id",
     "Honorer", "Guru Bahasa Indonesia", "active"],

    # Row 4: ERROR - Nama lengkap kosong (required field)
    ["198805052015042001", "2015123456700002", "", "male",
     "Jl. Merdeka No. 5", "081234567892", "budi.santoso@sekolah.sch.id",
     "PNS", "Guru IPA", "active"],

    # Row 5: ERROR - Jenis kelamin invalid
    ["198703032014033001", "2015123456700003", "Rina Wati, S.Pd", "laki-laki",
     "Jl. Sudirman No. 15", "081234567893", "rina.wati@sekolah.sch.id",
     "PNS", "Guru Seni", "active"],

    # Row 6: ERROR - Status invalid
    ["198602022013022001", "2015123456700004", "Dewi Lestari, M.Pd", "female",
     "Jl. Asia Afrika No. 20", "081234567894", "dewi.lestari@sekolah.sch.id",
     "PNS", "Guru IPS", "aktif"],

    # Row 7: ERROR - Email format invalid (no @)
    ["198901012016011001", "2015123456700005", "Eko Prasetyo, S.Kom", "male",
     "Jl. Gatot Subroto No. 25", "081234567895", "eko.prasetyo-sekolah.sch.id",
     "Honorer", "Guru TI", "active"],

    # Row 8: EMPTY ROW (will be skipped)
    ["", "", "", "", "", "", "", "", "", ""],

    # Row 9: ERROR - Duplicate NIP (same as row 2)
    ["198501012010011001", "2015123456700006", "Test Duplikat NIP", "male",
     "Jl. Test No. 1", "081234567896", "test.duplikatnip@sekolah.sch.id",
     "Honorer", "Guru Test", "active"],

    # Row 10: VALID - Another valid teacher
    ["199001012017011001", "2015123456700007", "Maya Putri, S.Pd", "female",
     "Jl. Pahlawan No. 30", "081234567897", "maya.putri@sekolah.sch.id",
     "PNS", "Guru PPKN", "active"],

    # Row 11: ERROR - Nama lengkap kosong + gender invalid (multiple errors)
    ["199101012018011001", "2015123456700008", "", "wanita",
     "Jl. Kartini No. 5", "081234567898", "test.multierror@sekolah.sch.id",
     "Honorer", "Guru Olahraga", "active"],

    # Row 12: VALID - Inactive status
    ["198401012009011001", "2015123456700009", "Agus Setiawan, S.Pd", "male",
     "Jl. Veteran No. 12", "081234567899", "agus.setiawan@sekolah.sch.id",
     "PNS", "Guru Penjaskes", "inactive"],

    # Row 13: ERROR - Duplicate email (same as row 3)
    ["199201012019011001", "2015123456700010", "Rudi Hermanto, S.Kom", "male",
     "Jl. Diponegoro No. 8", "081234567800", "siti.nurhaliza@sekolah.sch.id",
     "Honorer", "Guru Teknik", "active"],
]

for row_num, row_data in enumerate(test_data, 2):
    for col, value in enumerate(row_data, 1):
        ws.cell(row=row_num, column=col, value=value)

# Save file
output_path = "/Users/edwinmalihu/Documents/internal/SIAKAD/test_import_guru.xlsx"
wb.save(output_path)
print(f"File created: {output_path}")
print(f"Total rows: {len(test_data)} data rows + 1 header")
print()
print("Expected results:")
print("- Row 2: VALID (berhasil)")
print("- Row 3: VALID (berhasil)")
print("- Row 4: ERROR (nama lengkap kosong)")
print("- Row 5: ERROR (jenis kelamin invalid)")
print("- Row 6: ERROR (status invalid)")
print("- Row 7: ERROR (email format invalid)")
print("- Row 8: SKIPPED (empty row)")
print("- Row 9: ERROR (duplicate NIP)")
print("- Row 10: VALID (berhasil)")
print("- Row 11: ERROR (nama kosong + gender invalid)")
print("- Row 12: VALID (berhasil, inactive)")
print("- Row 13: ERROR (duplicate email)")
print()
print("Summary: 4 berhasil, 1 dilewati, 7 gagal")
